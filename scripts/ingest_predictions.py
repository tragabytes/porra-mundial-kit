"""Inyecta las predicciones de los jugadores (archivos en predictions/) en el ADMIN.

Convenciones:
- Filename: `<NN>_<nombre>.xlsx` donde NN es 01..25 (slot del jugador)
  Ejemplos válidos: `01_juan.xlsx`, `12_maria-perez.xlsx`
- Cada jugador rellena su copia de `Excel-Mundial-2026.xlsx` (template) con sus
  predicciones en el grid WORLDCUP. La hoja `Pool` del player ya formatea los
  datos listos para pegar (col C contiene "X|Y-Z" para cada partido).

Mapeo de celdas (verificado contra ADMIN-Excel-Mundial-2026.xlsx paid 25 jugadores):

| Destino (ADMIN)              | Origen (player)   | Filas en ambos lados    | Contenido        |
|------------------------------|-------------------|-------------------------|------------------|
| <slot_col>5                  | Home!C10          | R5                      | Nombre jugador   |
| <slot_col>6:<slot_col>29     | Pool!C6:C29       | 24 partidos             | Jornada 1        |
| <slot_col>30:<slot_col>53    | Pool!C30:C53      | 24 partidos             | Jornada 2        |
| <slot_col>54:<slot_col>77    | Pool!C54:C77      | 24 partidos             | Jornada 3        |
| <slot_col>80:<slot_col>127   | Pool!C80:C127     | 48 posiciones           | Pos. Grupos      |
| <slot_col>130:<slot_col>161  | Pool!C130:C161    | 32 clasificados         | Equipos 1/16     |
| <slot_col>164:<slot_col>179  | Pool!C164:C179    | 16 partidos             | Dieciseisavos    |
| <slot_col>182:<slot_col>197  | Pool!C182:C197    | 16 clasificados         | Equipos 1/8      |
| <slot_col>200:<slot_col>207  | Pool!C200:C207    | 8 partidos              | Octavos          |
| <slot_col>210:<slot_col>217  | Pool!C210:C217    | 8 clasificados          | Equipos 1/4      |
| <slot_col>220:<slot_col>223  | Pool!C220:C223    | 4 partidos              | Cuartos          |
| <slot_col>226:<slot_col>229  | Pool!C226:C229    | 4 clasificados          | Equipos 1/2      |
| <slot_col>232:<slot_col>233  | Pool!C232:C233    | 2 partidos              | Semifinales      |
| <slot_col>236:<slot_col>237  | Pool!C236:C237    | 2 clasificados          | Equipos 3-4      |
| <slot_col>240:<slot_col>241  | Pool!C240:C241    | 2 clasificados          | Equipos Final    |
| <slot_col>244:<slot_col>247  | Pool!C244:C247    | 4 filas (244+247 datos) | 3-4 + Final      |
| <slot_col>250:<slot_col>252  | Pool!C250:C252    | 3 (camp/sub/3º)         | Honor equipos    |
| <slot_col>253:<slot_col>258  | Pool!C253:C258    | 6 entradas              | Cuadro de Honor  |

IMPORTANTE (corregido 28/06): las predicciones de POSICIÓN DE GRUPOS y de EQUIPOS
CLASIFICADOS (qué selecciones pasan cada ronda) SÍ hay que copiarlas — el ADMIN NO
las deriva solo. El template del jugador ya las trae resueltas como NOMBRE de país en
Pool!C80:C127 (posiciones), Pool!C130:C161 / C182:C197 / ... (clasificados por ronda)
y Pool!C250:C252 (campeón/subcampeón/3º). Sin estos bloques, las celdas COUNTIF del
ADMIN que puntúan posiciones/clasificados quedan a 0 (bug detectado en la auditoría
post-grupos del Mundial 2026).

Slot → columna en ADMIN: cada jugador ocupa 3 cols contiguas.
- Slot 1 → S (col 19), 2 → V (22), ..., 25 → CM (91)
- Fórmula: first_col = 19 + (slot - 1) * 3

Usage:
  python scripts/ingest_predictions.py
  python scripts/ingest_predictions.py --admin path/to/ADMIN.xlsx --predictions path/to/dir/
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Importamos desde el mismo directorio
sys.path.insert(0, str(Path(__file__).parent))
from lib_excel import load, save, unprotect_all_sheets, reprotect_all_sheets, recalc

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_ADMIN = PROJECT_ROOT / "ADMIN-Excel-Mundial-2026.xlsx"
DEFAULT_PREDICTIONS_DIR = PROJECT_ROOT / "predictions"

# Mapeo slot → columna ---
FIRST_SLOT_COL = 19  # S

def slot_to_first_col(slot: int) -> int:
    """Slot 1..25 → primera columna (1-indexed) del jugador en ADMIN.
    Slot 1 → 19 (S), 2 → 22 (V), 3 → 25 (Y), ..., 25 → 91 (CM).
    """
    if not 1 <= slot <= 25:
        raise ValueError(f"Slot fuera de rango: {slot} (debe ser 1..25)")
    return FIRST_SLOT_COL + (slot - 1) * 3

# Rangos de copy-paste (Pool!C{start}:C{end} → ADMIN!<slot_col>{start}:{end}) ---
# Coincidencia 1-a-1 de fila porque matejero diseñó Pool con la misma indexación.
PASTE_RANGES: list[tuple[int, int, str]] = [
    (6, 29, "Jornada 1"),
    (30, 53, "Jornada 2"),
    (54, 77, "Jornada 3"),
    (80, 127, "Pos. Grupos"),          # posiciones exactas de cada grupo (nombre de país)
    (130, 161, "Equipos 1/16"),        # 32 clasificados a dieciseisavos
    (164, 179, "Dieciseisavos"),
    (182, 197, "Equipos 1/8"),         # 16 clasificados a octavos
    (200, 207, "Octavos"),
    (210, 217, "Equipos 1/4"),         # 8 clasificados a cuartos
    (220, 223, "Cuartos"),
    (226, 229, "Equipos 1/2"),         # 4 clasificados a semis
    (232, 233, "Semifinales"),
    (236, 237, "Equipos 3-4"),         # finalistas del 3-4 puesto
    (240, 241, "Equipos Final"),       # finalistas de la final
    (244, 247, "3-4 puesto + Final"),  # rows 245-246 son vacíos / separadores
    (250, 252, "Honor equipos"),       # campeón/subcampeón/3º (nombre de país)
    (253, 258, "Cuadro de Honor"),
]
POOL_COL = 3  # C en Pool
NAME_ROW = 5
PLAYER_NAME_CELL = ("Home", "C10")

# Marcador "0|-" = predicción no-rellenada por defecto en Pool del player template
DEFAULT_PREDICTION_MARKER = "0|-"

# Parsing del filename ---
FILENAME_RE = re.compile(r"^(\d{1,2})[_\- ](.+)\.xlsx$", re.IGNORECASE)

def parse_slot_from_filename(path: Path) -> tuple[int, str]:
    """`01_juan.xlsx` → (1, 'juan'). Lanza si formato inválido."""
    m = FILENAME_RE.match(path.name)
    if not m:
        raise ValueError(
            f"Filename no respeta el formato 'NN_nombre.xlsx': {path.name}"
        )
    return int(m.group(1)), m.group(2)

# Lógica principal ---

def read_player_predictions(player_xlsx: Path) -> dict:
    """Lee del player workbook (data_only=True) el nombre y todos los rangos."""
    wb = load_workbook(str(player_xlsx), data_only=True)
    sheet_name, cell = PLAYER_NAME_CELL
    name = wb[sheet_name][cell].value
    if not name or str(name).strip() in ("", "Nombre"):
        name = None  # caer al label del filename

    pool = wb["Pool"]
    ranges: dict[str, list] = {}
    for start, end, label in PASTE_RANGES:
        values = [pool.cell(row=r, column=POOL_COL).value for r in range(start, end + 1)]
        ranges[label] = values
    return {"name": name, "ranges": ranges}


def inject_into_admin(
    admin_wb,
    slot: int,
    fallback_name: str,
    predictions: dict,
) -> None:
    """Escribe nombre + todos los rangos del jugador `slot` en la hoja ADMIN."""
    ws = admin_wb["ADMIN"]
    col = slot_to_first_col(slot)
    col_letter = get_column_letter(col)

    name = predictions.get("name") or fallback_name
    ws.cell(row=NAME_ROW, column=col).value = name

    non_default = 0
    for (start, end, label), values in zip(PASTE_RANGES, predictions["ranges"].values()):
        for offset, value in enumerate(values):
            target_row = start + offset
            ws.cell(row=target_row, column=col).value = value
            if value not in (None, DEFAULT_PREDICTION_MARKER):
                non_default += 1

    print(f"  Slot {slot:>2} ({col_letter}): nombre='{name}', "
          f"{non_default} predicciones no-default")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--admin", type=Path, default=DEFAULT_ADMIN,
                        help="Ruta al ADMIN-Excel-Mundial-2026.xlsx")
    parser.add_argument("--predictions", type=Path, default=DEFAULT_PREDICTIONS_DIR,
                        help="Directorio con los .xlsx de jugadores")
    args = parser.parse_args()

    if not args.admin.exists():
        print(f"ERROR: no encuentro ADMIN en {args.admin}", file=sys.stderr)
        return 1
    if not args.predictions.exists():
        print(f"ERROR: no encuentro directorio en {args.predictions}", file=sys.stderr)
        return 1

    player_files = sorted(args.predictions.glob("*.xlsx"))
    if not player_files:
        print(f"No hay .xlsx en {args.predictions}. Nada que hacer.")
        return 0

    print(f"Cargando ADMIN: {args.admin}")
    admin_wb = load(args.admin)
    unprotect_all_sheets(admin_wb)

    seen_slots: dict[int, Path] = {}
    for p in player_files:
        try:
            slot, _ = parse_slot_from_filename(p)
        except ValueError as e:
            print(f"  SKIP {p.name}: {e}", file=sys.stderr)
            continue
        if slot in seen_slots:
            print(f"ERROR: slot {slot} duplicado entre {seen_slots[slot].name} y {p.name}",
                  file=sys.stderr)
            return 2
        seen_slots[slot] = p

    print(f"Procesando {len(seen_slots)} archivos de predicción...")
    for slot, p in sorted(seen_slots.items()):
        try:
            preds = read_player_predictions(p)
        except Exception as e:
            print(f"  ERROR leyendo {p.name}: {e}", file=sys.stderr)
            return 3
        _, label = parse_slot_from_filename(p)
        inject_into_admin(admin_wb, slot, fallback_name=label, predictions=preds)

    reprotect_all_sheets(admin_wb)
    print(f"Guardando ADMIN: {args.admin}")
    save(admin_wb, args.admin)
    try:
        print("Recalculando con LibreOffice headless...")
        recalc(args.admin)
        print("OK. Recálculo completado: CLAS/Stats/Daily quedan con valores frescos.")
    except Exception as e:
        print(f"AVISO: predicciones guardadas, pero no se pudo recalcular ({e}). "
              f"Ábrelo en Excel y pulsa Ctrl+Alt+F9, o instala LibreOffice.", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
