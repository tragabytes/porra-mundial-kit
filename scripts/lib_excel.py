"""Helpers básicos para manipular los Excel de matejero.

Responsabilidades:
- Cargar/guardar workbooks con openpyxl.
- Desproteger/reproteger hojas (matejero las protege para que el usuario no rompa fórmulas).
- Forzar recálculo invocando LibreOffice headless (openpyxl no ejecuta fórmulas).
"""

from __future__ import annotations

import re
import shutil
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def load(path: str | Path) -> Workbook:
    """Carga un workbook conservando fórmulas (data_only=False)."""
    return load_workbook(str(path), data_only=False)


def load_values(path: str | Path) -> Workbook:
    """Carga un workbook leyendo los valores cacheados (data_only=True).

    Importante: los valores son los que LibreOffice/Excel cachearon en su último
    guardado. Si has modificado el archivo con openpyxl sin recalcular, estos
    valores estarán desactualizados.
    """
    return load_workbook(str(path), data_only=True)


def unprotect_all_sheets(wb: Workbook) -> None:
    """Desprotege todas las hojas del workbook (incluidas las ocultas)."""
    for ws in wb.worksheets:
        ws.protection.sheet = False


def reprotect_all_sheets(wb: Workbook) -> None:
    """Reproteje todas las hojas. Sin password (suficiente para evitar ediciones accidentales en Excel UI)."""
    for ws in wb.worksheets:
        ws.protection.sheet = True


def save(wb: Workbook, path: str | Path) -> None:
    """Guarda el workbook en disco."""
    wb.save(str(path))


def _find_soffice() -> str:
    """Localiza el ejecutable de LibreOffice/soffice.

    Orden: PATH > rutas conocidas de Windows. Lanza excepción si no encuentra.
    """
    found = shutil.which("soffice") or shutil.which("libreoffice")
    if found:
        return found
    # Fallbacks Windows habituales
    candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for c in candidates:
        if Path(c).exists():
            return c
    raise RuntimeError(
        "No encuentro LibreOffice/soffice. Instálalo o añádelo al PATH."
    )


def recalc(xlsx_path: str | Path, timeout: int = 300) -> None:
    """Fuerza el recálculo de todas las fórmulas del xlsx in-place.

    Mecanismo: LibreOffice abre el archivo en modo headless (lo cual ya dispara
    el recálculo) y lo reexporta a .xlsx. Sustituimos el original por el
    reexportado.

    Usa un perfil aislado por ejecución para evitar conflictos con instancias
    de LibreOffice abiertas en interactivo.
    """
    xlsx_path = Path(xlsx_path).resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)

    soffice = _find_soffice()

    with tempfile.TemporaryDirectory(prefix="porra_recalc_") as tmp:
        out_dir = Path(tmp) / "out"
        out_dir.mkdir()
        profile_dir = Path(tmp) / "profile"
        profile_dir.mkdir()
        profile_url = profile_dir.as_uri()

        cmd = [
            soffice,
            f"-env:UserInstallation={profile_url}",
            "--headless",
            "--calc",
            "--convert-to", "xlsx",
            "--outdir", str(out_dir),
            str(xlsx_path),
        ]
        result = subprocess.run(
            cmd, timeout=timeout, capture_output=True, text=True
        )
        if result.returncode != 0:
            raise RuntimeError(
                f"LibreOffice recalc falló (exit {result.returncode}):\n"
                f"STDOUT: {result.stdout}\nSTDERR: {result.stderr}"
            )

        recalculated = out_dir / xlsx_path.name
        if not recalculated.exists():
            raise RuntimeError(
                f"LibreOffice no produjo {recalculated}. STDOUT: {result.stdout}"
            )
        shutil.move(str(recalculated), str(xlsx_path))


# ---------------------------------------------------------------------------
# Lectura de predicciones y datos "del día" (V7)
#
# El bot cita qué predijo cada jugador para un partido y el ranking de la
# jornada. Las predicciones viven en la hoja ADMIN; el ranking del día lo
# calcula matejero en la hoja DailyClas (depende del selector DailyPrediction!H1).
# ---------------------------------------------------------------------------

# Cada jugador ocupa 3 columnas en ADMIN; la 1ª es 19 + (slot-1)*3 (= S, V, Y…).
# Misma fórmula que slot_to_first_col en ingest_predictions.py (se replica para
# no crear un import cruzado entre scripts hermanos).
_FIRST_SLOT_COL = 19
_MAX_SLOTS = 25
_NAME_ROW = 5
_ADMIN_MATCH_COL = 11  # K: nombre del partido (fórmula que referencia WORLDCUP)
_ADMIN_DATE_COL = 8    # H: fecha del partido (fórmula; valor cacheado = datetime)

# La columna K del ADMIN referencia, por cada fila de partido, la fila de
# WORLDCUP correspondiente, p.ej. "=CONCATENATE(WORLDCUP!AA4,...)" (grupos) o
# "=WORLDCUP!W101" (eliminatorias). Leyendo la fórmula en crudo reconstruimos
# el mapeo fila_WORLDCUP -> fila_ADMIN sin necesidad de recalcular.
_WORLDCUP_REF_RE = re.compile(r"WORLDCUP!\$?[A-Z]+\$?(\d+)")

# Predicción de un jugador: "signo|local-visitante", p.ej. "1|2-1", "X|1-1".
# En eliminatorias la celda trae basura delante (nombres + carácter de control),
# p.ej. "República Checa-Canadá\x96X|1-1"; por eso tomamos el ÚLTIMO token.
_PREDICTION_RE = re.compile(r"([12X])\|(\d+)-(\d+)")


def build_admin_row_map(path: str | Path) -> dict[int, int]:
    """Devuelve {fila_WORLDCUP: fila_ADMIN} leyendo las fórmulas de ADMIN!K.

    Solo incluye filas cuya fórmula referencia a WORLDCUP; las demás
    (separadores, cabeceras "Idiomas!…") no matchean y se descartan solas.
    """
    ws = load(path)["ADMIN"]
    row_map: dict[int, int] = {}
    for r in range(6, 259):
        formula = ws.cell(row=r, column=_ADMIN_MATCH_COL).value
        if not isinstance(formula, str):
            continue
        m = _WORLDCUP_REF_RE.search(formula)
        if m:
            row_map[int(m.group(1))] = r
    return row_map


def parse_prediction(value) -> tuple[str, int, int] | None:
    """'1|2-1' -> ('1', 2, 1). None si no hay predicción válida.

    Cubre None, '-', 'Pegar Valores…' (devuelven None) y la basura delante de
    las predicciones de eliminatorias (toma el último token 'signo|m-n').
    """
    if not isinstance(value, str):
        return None
    matches = _PREDICTION_RE.findall(value)
    if not matches:
        return None
    sign, home, away = matches[-1]
    return sign, int(home), int(away)


def read_matches_predictions(path: str | Path,
                             admin_rows: list[int]) -> dict[int, dict[str, tuple[str, int, int]]]:
    """{fila_ADMIN: {nombre: (signo, local, visitante)}} para varios partidos.

    Carga el workbook UNA sola vez (700 KB por carga: leer N partidos con la
    versión de fila única costaba N cargas). Lee valores cacheados
    (data_only=True), salta slots sin cargar ('Pegar Valores…') y predicciones
    no parseables.
    """
    ws = load_values(path)["ADMIN"]
    out: dict[int, dict[str, tuple[str, int, int]]] = {r: {} for r in admin_rows}
    for slot in range(1, _MAX_SLOTS + 1):
        col = _FIRST_SLOT_COL + (slot - 1) * 3
        name = ws.cell(row=_NAME_ROW, column=col).value
        if not name or (isinstance(name, str) and name.startswith("Pegar")):
            continue
        for admin_row in admin_rows:
            pred = parse_prediction(ws.cell(row=admin_row, column=col).value)
            if pred is not None:
                out[admin_row][str(name)] = pred
    return out


def read_match_predictions(path: str | Path, admin_row: int) -> dict[str, tuple[str, int, int]]:
    """{nombre_jugador: (signo, local, visitante)} para un partido (fila ADMIN)."""
    return read_matches_predictions(path, [admin_row])[admin_row]


def read_all_match_predictions(path: str | Path, admin_map: dict[int, int],
                               home_col: int, away_col: int,
                               home_score_col: int | None = None,
                               away_score_col: int | None = None) -> list[dict]:
    """Picks de TODOS los partidos del torneo, en una sola carga del workbook.

    `admin_map` es el {fila_WORLDCUP: fila_ADMIN} de build_admin_row_map();
    home_col/away_col son las columnas de equipos en WORLDCUP (las define el
    caller). Devuelve, ordenado por fila WORLDCUP:
      [{row, home, away, fecha, predicciones}]
    Las eliminatorias sin resolver llevan home/away None (el caller filtra);
    fecha es datetime o None (eliminatorias).

    Si se pasan home_score_col/away_score_col (columnas de goles en WORLDCUP),
    cada entrada añade home_score/away_score (int si el partido se ha jugado,
    None si la celda está vacía). Sirve para saber qué partidos tienen resultado.
    """
    wb = load_values(path)
    admin = wb["ADMIN"]
    wc = wb["WORLDCUP"]

    name_cols: dict[int, str] = {}
    for slot in range(1, _MAX_SLOTS + 1):
        col = _FIRST_SLOT_COL + (slot - 1) * 3
        name = admin.cell(row=_NAME_ROW, column=col).value
        if name and not (isinstance(name, str) and name.startswith("Pegar")):
            name_cols[col] = str(name)

    out: list[dict] = []
    for wrow in sorted(admin_map):
        arow = admin_map[wrow]
        h = wc.cell(row=wrow, column=home_col).value
        a = wc.cell(row=wrow, column=away_col).value
        fecha = admin.cell(row=arow, column=_ADMIN_DATE_COL).value
        preds = {}
        for col, name in name_cols.items():
            p = parse_prediction(admin.cell(row=arow, column=col).value)
            if p is not None:
                preds[name] = p
        entry = {
            "row": wrow,
            "home": h.strip() if isinstance(h, str) and h.strip() else None,
            "away": a.strip() if isinstance(a, str) and a.strip() else None,
            "fecha": fecha if isinstance(fecha, datetime) else None,
            "predicciones": preds,
        }
        if home_score_col is not None and away_score_col is not None:
            hs = wc.cell(row=wrow, column=home_score_col).value
            as_ = wc.cell(row=wrow, column=away_score_col).value
            entry["home_score"] = int(hs) if isinstance(hs, (int, float)) else None
            entry["away_score"] = int(as_) if isinstance(as_, (int, float)) else None
        out.append(entry)
    return out


def read_match_date(path: str | Path, admin_row: int) -> datetime | None:
    """Fecha del partido (ADMIN!H<row>) o None si no es fecha (eliminatorias).

    H es una fórmula; leemos su valor cacheado. En fase de grupos da la fecha
    del partido; en eliminatorias da un label de texto ("Dieciseisavos…").
    """
    v = load_values(path)["ADMIN"].cell(row=admin_row, column=_ADMIN_DATE_COL).value
    return v if isinstance(v, datetime) else None


def read_daily_ranking(path: str | Path) -> list[dict]:
    """Ranking de la jornada desde la hoja DailyClas. Asume recálculo previo.

    Requiere que DailyPrediction!H1 apunte al día deseado antes del recalc.
    Lee nombre (col F) y puntos del día (col G) de las filas 4-28, salta los
    slots sin cargar y ordena por puntos desc (la posición de la propia hoja no
    es fiable con empates). Devuelve [{name, points, position}].
    """
    ws = load_values(path)["DailyClas"]
    rows = []
    for r in range(4, 29):
        name = ws.cell(row=r, column=6).value
        if not name or (isinstance(name, str) and name.startswith("Pegar")):
            continue
        rows.append({
            "name": str(name),
            "points": int(ws.cell(row=r, column=7).value or 0),
        })
    rows.sort(key=lambda x: -x["points"])
    for i, row in enumerate(rows, start=1):
        row["position"] = i
    return rows
