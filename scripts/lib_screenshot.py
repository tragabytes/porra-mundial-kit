"""Renderiza el leaderboard como PNG via Playwright + Jinja2.

Carga el template `design/leaderboard.html`, lo rellena con datos del partido
y del ranking, y lanza Chromium headless para hacer un screenshot 1080x1920
listo para WhatsApp.

Por qué Playwright y no Pillow: la plantilla del diseñador usa Google Fonts
(Oswald/IBM Plex), tramas halftone (radial-gradient), gradientes y posicionado
complejo. Reproducir eso fielmente en Pillow son cientos de líneas y se pierde
fidelidad al diseño.

Usage:
  from lib_screenshot import render_leaderboard
  png = render_leaderboard(
      "ADMIN.xlsx",
      match=match_dict,           # de lib_football_api.fetch_finished_matches()
      home_es="España", away_es="Cabo Verde",
      ranking_before=[...],       # opcional: para el titular (cambio de líder)
      next_match="México vs Sudáfrica · 11/06 19:00",  # opcional
      matches_remaining=103,      # opcional
  )

  python scripts/lib_screenshot.py [ADMIN.xlsx]   # smoke test
"""

from __future__ import annotations

import base64
import json
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DESIGN_DIR = PROJECT_ROOT / "design"
TEMPLATE_NAME = "leaderboard.html"
TEAMS_ISO_PATH = PROJECT_ROOT / "data" / "teams_iso.json"
QUOTES_PATH = PROJECT_ROOT / "data" / "quotes_futbol.json"
LOGO_PATH = DESIGN_DIR / "assets" / "wc26_logo.svg"

# WC 2026: 12 grupos × 6 + 32 eliminatorias = 72 + 32 = 104 partidos
WC_TOTAL_MATCHES = 104

# Etiquetas de fase a partir de football-data.org `stage`.
# GROUP_STAGE se compone con grupo y matchday (no está aquí).
STAGE_LABELS = {
    "LAST_32": "Dieciseisavos de final",
    "LAST_16": "Octavos de final",
    "QUARTER_FINALS": "Cuartos de final",
    "SEMI_FINALS": "Semifinales",
    "THIRD_PLACE": "Tercer y cuarto puesto",
    "FINAL": "Final",
}


def _load_teams_iso() -> dict[str, str]:
    raw = json.loads(TEAMS_ISO_PATH.read_text(encoding="utf-8"))
    return {k: v for k, v in raw.items() if not k.startswith("_")}


def _fase_label(stage: str, group: str | None, matchday: int | None) -> str:
    if stage == "GROUP_STAGE":
        if group and matchday:
            return f"Grupo {group} · Jornada {matchday}"
        return "Fase de grupos"
    return STAGE_LABELS.get(stage, stage or "—")


def _estado_label(duration: str, home_pens: int | None, away_pens: int | None) -> str:
    if duration == "PENALTY_SHOOTOUT" and home_pens is not None and away_pens is not None:
        return f"Penales {home_pens}-{away_pens}"
    if duration == "EXTRA_TIME":
        return "AET"
    return "FT"


def _read_ranking_after(admin_xlsx: Path | str) -> list[dict]:
    """Lee CLAS post-recálculo. Devuelve [{pos, name, points}] ordenado."""
    wb = load_workbook(str(admin_xlsx), data_only=True)
    ws = wb["CLAS"]
    n_players = int(wb["ADMIN"]["D5"].value or 25)
    rows = []
    for slot in range(1, n_players + 1):
        r = 4 + slot
        name = ws.cell(row=r, column=3).value
        if not name or (isinstance(name, str) and name.startswith("Pegar")):
            continue
        rows.append({
            "name": str(name),
            "points": int(ws.cell(row=r, column=4).value or 0),
        })
    rows.sort(key=lambda x: (-x["points"], x["name"].lower()))
    for i, row in enumerate(rows, start=1):
        row["pos"] = i
    return rows


def _quote_of_the_day() -> tuple[str, str]:
    """(texto, autor) de data/quotes_futbol.json, rotando una por día.

    Índice determinista por fecha de Madrid: misma cita todo el día (y en
    ambas porras), sin repetir hasta agotar el ciclo. Clamp defensivo por si
    alguna cita larga se cuela en la BD: el lienzo nunca debe desbordar.
    """
    quotes = json.loads(QUOTES_PATH.read_text(encoding="utf-8"))
    today = datetime.now(ZoneInfo("Europe/Madrid")).date()
    q = quotes[today.toordinal() % len(quotes)]
    texto = str(q["texto"]).strip()
    if len(texto) > 150:
        texto = texto[:149].rstrip() + "…"
    return texto, str(q["autor"]).strip()


def _quote_font_size(texto: str) -> int:
    """Tamaño del texto de la cita según longitud, para que quepa siempre."""
    if len(texto) <= 90:
        return 30
    if len(texto) <= 120:
        return 26
    return 22


def _logo_data_uri() -> str:
    """Emblema del Mundial como data-URI (el render usa set_content: las
    rutas relativas a archivo no resuelven)."""
    b64 = base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii")
    return f"data:image/svg+xml;base64,{b64}"


def _build_titular(score_str: str,
                   ranking_before: list[dict],
                   ranking_after: list[dict]) -> str:
    """Heurística para el banner amarillo. Determinista, sin LLM.

    - Cambio de líder → "X ARREBATA EL LIDERATO A Y"
    - Resto         → "CLASIFICACIÓN ACTUALIZADA · MARCADOR"
    """
    if ranking_before and ranking_after:
        leader_before = ranking_before[0]["name"]
        leader_after = ranking_after[0]["name"]
        if leader_before != leader_after:
            return f"{leader_after.upper()} ARREBATA EL LIDERATO A {leader_before.upper()}"
    return f"CLASIFICACIÓN ACTUALIZADA · {score_str.upper()}"


def _build_template_data(
    *,
    match: dict | None,
    home_es: str | None,
    away_es: str | None,
    ranking_before: list[dict] | None,
    ranking_after: list[dict],
    titular: str | None,
    next_match: str | None,
    matches_remaining: int | None,
    teams_iso: dict[str, str],
    fase_label: str | None = None,
) -> dict:
    # Modo "edición diaria" (match=None): sin bloque de resultado; la cabecera
    # lleva fase_label y el titular lo aporta el caller (no hay marcador del
    # que derivarlo).
    if match is None:
        fase = fase_label or "Clasificación general"
        titular_final = titular or "ASÍ VA LA PORRA"
        edicion = "Edición diaria"
        partido = None
    else:
        fase = _fase_label(
            match.get("stage", "GROUP_STAGE"),
            match.get("group"),
            match.get("matchday"),
        )
        estado = _estado_label(
            match.get("duration", "REGULAR"),
            match.get("home_penalties"),
            match.get("away_penalties"),
        )

        score_str = f"{home_es} {match['home_score']}-{match['away_score']} {away_es}"
        titular_final = titular or _build_titular(
            score_str, ranking_before or [], ranking_after,
        )
        edicion = "Edición tras partido"
        partido = {
            "local": {
                "nombre": home_es.upper(),
                "iso": teams_iso.get(home_es, "???"),
                "goles": match["home_score"],
            },
            "visitante": {
                "nombre": away_es.upper(),
                "iso": teams_iso.get(away_es, "???"),
                "goles": match["away_score"],
            },
            "estado": estado,
        }

    enriched = [
        {"pos": r["pos"], "nombre": r["name"], "puntos": r["points"]}
        for r in ranking_after
    ]

    podio = enriched[:3]
    while len(podio) < 3:
        podio.append({"pos": len(podio) + 1, "nombre": "—", "puntos": 0})

    lista_4_N = enriched[3:25]  # diseño deja sitio para 22 filas
    ultima_posicion = enriched[-1]["pos"] if enriched else 0

    cita, cita_autor = _quote_of_the_day()
    return {
        "fase": fase,
        "edicion": edicion,
        "partido": partido,
        "titular_delta": titular_final,
        "podio": podio,
        "lista_4_N": lista_4_N,
        "ultima_posicion": ultima_posicion,
        "proximo": next_match or "—",
        "partidos_restantes": matches_remaining if matches_remaining is not None else "?",
        "cita": cita,
        "cita_autor": cita_autor,
        "cita_size": _quote_font_size(cita),
        "logo_b64": _logo_data_uri(),
    }


def render_html(
    *,
    admin_xlsx: Path | str,
    match: dict | None = None,
    home_es: str | None = None,
    away_es: str | None = None,
    ranking_before: list[dict] | None = None,
    titular: str | None = None,
    next_match: str | None = None,
    matches_remaining: int | None = None,
    fase_label: str | None = None,
) -> str:
    """Devuelve el HTML renderizado (sin lanzar Chromium). Útil para preview.

    Con match=None renderiza la "edición diaria": solo clasificación, sin
    bloque de resultado (usado por send_daily_leaderboard.py).
    """
    teams_iso = _load_teams_iso()
    ranking_after = _read_ranking_after(admin_xlsx)
    data = _build_template_data(
        match=match, home_es=home_es, away_es=away_es,
        ranking_before=ranking_before, ranking_after=ranking_after,
        titular=titular, next_match=next_match,
        matches_remaining=matches_remaining,
        teams_iso=teams_iso, fase_label=fase_label,
    )
    env = Environment(
        loader=FileSystemLoader(str(DESIGN_DIR)),
        autoescape=select_autoescape(["html", "xml"]),
    )
    template = env.get_template(TEMPLATE_NAME)
    return template.render(**data)


def render_leaderboard(
    admin_xlsx: Path | str,
    *,
    match: dict | None = None,
    home_es: str | None = None,
    away_es: str | None = None,
    ranking_before: list[dict] | None = None,
    titular: str | None = None,
    next_match: str | None = None,
    matches_remaining: int | None = None,
    fase_label: str | None = None,
) -> bytes:
    """Renderiza el leaderboard como PNG 1080x1920 (bytes).

    Con match=None genera la "edición diaria" (solo clasificación).
    """
    html = render_html(
        admin_xlsx=admin_xlsx, match=match,
        home_es=home_es, away_es=away_es,
        ranking_before=ranking_before, titular=titular,
        next_match=next_match, matches_remaining=matches_remaining,
        fase_label=fase_label,
    )

    # Import perezoso: Playwright pesa, no debería importarse en tests sin render.
    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch()
        try:
            ctx = browser.new_context(
                viewport={"width": 1080, "height": 1920},
                device_scale_factor=1,
            )
            page = ctx.new_page()
            page.set_content(html, wait_until="networkidle")
            # Asegura que las Google Fonts terminen de cargar antes del snapshot
            page.wait_for_function("document.fonts.ready")
            png_bytes = page.screenshot(type="png", full_page=False)
        finally:
            browser.close()
    return png_bytes


if __name__ == "__main__":
    admin = sys.argv[1] if len(sys.argv) > 1 else "ADMIN-Excel-Mundial-2026.xlsx"
    fake_match = {
        "home": "Spain", "away": "Cape Verde",
        "home_score": 2, "away_score": 1,
        "duration": "REGULAR",
        "home_penalties": None, "away_penalties": None,
        "stage": "GROUP_STAGE", "group": "H", "matchday": 1,
    }
    png = render_leaderboard(
        admin, match=fake_match,
        home_es="España", away_es="Cabo Verde",
        ranking_before=None, titular=None,
        next_match="México vs Sudáfrica · 11/06 · 19:00",
        matches_remaining=103,
    )
    out = Path("leaderboard.png")
    out.write_bytes(png)
    print(f"PNG escrito en {out.resolve()} ({len(png)} bytes)")
