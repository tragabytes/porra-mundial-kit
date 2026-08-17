"""One-shot: genera data/match_row_map.json a partir del WORLDCUP del ADMIN.

Mapea cada partido (par de equipos en español) a su fila en la hoja WORLDCUP
del ADMIN. El cron lo usa para localizar a qué fila escribir los goles oficiales
cuando termina un partido.

Solo mapea la **fase de grupos** (72 partidos). Las eliminatorias tienen
placeholders ("1A", "2B", "W73", "L101") hasta que se resuelve la fase de
grupos; el cron las resolverá en tiempo real leyendo el WORLDCUP fresh.

Layout descubierto del WORLDCUP del ADMIN paid (matejero 25 jugadores):
- col 27 (AA): nombre equipo local (en español)
- col 32 (AF): nombre equipo visitante (en español)
- col 34 (AH): número de partido (1-104, numeración interna de matejero)
- col 29 (AC): goles local (input que el cron escribe)
- col 30 (AD): goles visitante (input que el cron escribe)

Rows fase grupos: 4-9 (A), 12-17 (B), 20-25 (C), 28-33 (D), 36-41 (E),
44-49 (F), 52-57 (G), 60-65 (H), 68-73 (I), 76-81 (J), 84-89 (K), 92-97 (L).

Usage:
  python scripts/bootstrap_match_rows.py
"""

from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
ADMIN_PATH = PROJECT_ROOT / "ADMIN-Excel-Mundial-2026.xlsx"
TEAMS_JSON = PROJECT_ROOT / "data" / "teams_en_es.json"
OUT_PATH = PROJECT_ROOT / "data" / "match_row_map.json"

COL_HOME = 27  # AA
COL_AWAY = 32  # AF
COL_MATCH_NUM = 34  # AH


def load_known_teams_es() -> set[str]:
    """Devuelve el set de nombres de equipo en español según teams_en_es.json."""
    data = json.loads(TEAMS_JSON.read_text(encoding="utf-8"))
    return {v for k, v in data.items() if not k.startswith("_")}


def main() -> int:
    if not ADMIN_PATH.exists():
        print(f"ERROR: no encuentro {ADMIN_PATH}", file=sys.stderr)
        return 1

    known_teams = load_known_teams_es()
    print(f"Equipos conocidos en teams_en_es.json: {len(known_teams)}")

    wb = load_workbook(str(ADMIN_PATH), data_only=True)
    ws = wb["WORLDCUP"]

    mapping: dict[str, int] = {}
    duplicates: list[tuple[str, int, int]] = []

    for r in range(4, ws.max_row + 1):
        home = ws.cell(row=r, column=COL_HOME).value
        away = ws.cell(row=r, column=COL_AWAY).value
        if not (isinstance(home, str) and isinstance(away, str)):
            continue
        if home not in known_teams or away not in known_teams:
            continue  # fila de eliminatorias con placeholder, la saltamos
        key = f"{home}|{away}"
        if key in mapping:
            duplicates.append((key, mapping[key], r))
            continue
        mapping[key] = r

    if duplicates:
        print("WARN: claves duplicadas detectadas (esto no debería pasar en fase grupos):",
              file=sys.stderr)
        for k, r1, r2 in duplicates:
            print(f"  {k}: rows {r1} y {r2}", file=sys.stderr)

    if len(mapping) != 72:
        print(f"WARN: esperaba 72 partidos de fase de grupos, encontré {len(mapping)}",
              file=sys.stderr)

    output = {
        "_comment": (
            "Map (equipo_local|equipo_visitante) → fila en WORLDCUP del ADMIN. "
            "Solo fase de grupos. Para eliminatorias, ingest_match_results.py "
            "consulta WORLDCUP en runtime tras resolución de grupos."
        ),
        "_generated_at": datetime.now(timezone.utc).isoformat(),
        "_source": ADMIN_PATH.name,
        "_total": len(mapping),
        "matches": dict(sorted(mapping.items())),
    }

    OUT_PATH.write_text(
        json.dumps(output, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"OK. {len(mapping)} partidos de fase grupos mapeados a {OUT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
