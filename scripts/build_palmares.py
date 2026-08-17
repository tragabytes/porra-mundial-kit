"""Destila el histórico de una edición cerrada en un palmarés por jugador.

Uso (defaults apuntan al archivo de la edición 2026):
    python scripts/build_palmares.py
    python scripts/build_palmares.py --pools-dir editions/mundial-2026/pools \
        --vps-dir editions/mundial-2026/vps --out-dir editions/mundial-2026

Genera `palmares.json` (para que el bot de la siguiente edición lo consuma) y
`palmares.md` (legible, para validación humana). Solo lectura de los datos de
la edición; asume los ADMIN recalculados (valores cacheados, sin LibreOffice).

Falla ruidosamente (assert) si el desglose de CLAS no suma el total de cada
jugador o si el ranking derivado no coincide con el leaderboard del state.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from lib_scoring import CLAS_CATEGORIES, read_clas_breakdown, read_ranking  # noqa: E402
from lib_scoring import result_sign  # noqa: E402

PROJECT_ROOT = Path(__file__).parent.parent

# Bloque Cuadro de Honor del ADMIN: fila -> clave del palmarés.
HONOR_ROWS = {
    250: "campeon", 251: "subcampeon", 252: "tercero",
    253: "bota_oro", 254: "bota_plata", 255: "bota_bronce",
    256: "balon_oro", 257: "balon_plata", 258: "balon_bronce",
}

# Los datos propios de la edición (premiados reales y notas) viven en un JSON
# junto a los datos archivados: {"real": {...}, "notas": [...]}. Se pasa con
# --meta; así el script es genérico y no lleva nombres de ninguna edición.


def _read_honor_block(admin_path: Path, slot_by_name: dict[str, int]) -> dict[str, dict]:
    """{nombre: {clave_premio: {"pick": str|None, "puntos": int}}} del bloque 250-258."""
    import warnings

    import openpyxl
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(admin_path, data_only=True)
    adm = wb["ADMIN"]
    out: dict[str, dict] = {}
    for name, slot in slot_by_name.items():
        col = 19 + (slot - 1) * 3
        premios = {}
        for row, clave in HONOR_ROWS.items():
            pick = adm.cell(row=row, column=col).value
            pts = adm.cell(row=row, column=col + 1).value
            premios[clave] = {"pick": pick if isinstance(pick, str) else None,
                              "puntos": int(pts) if isinstance(pts, (int, float)) else 0}
        out[name] = premios
    wb.close()
    return out


def _exactos_signos(all_predictions: list[dict]) -> tuple[dict, dict]:
    """Cuenta exactos y signos por jugador sobre los partidos con resultado."""
    ex, sg = {}, {}
    for entry in all_predictions:
        res = entry.get("resultado")
        if not res or "-" not in str(res):
            continue
        try:
            rh, ra = (int(x) for x in str(res).split("-"))
        except ValueError:
            continue
        for name, pick in (entry.get("predicciones") or {}).items():
            try:
                h, a = (int(x) for x in str(pick).split("-"))
            except (ValueError, AttributeError):
                continue
            if (h, a) == (rh, ra):
                ex[name] = ex.get(name, 0) + 1
            elif result_sign(h, a) == result_sign(rh, ra):
                sg[name] = sg.get(name, 0) + 1
    return ex, sg


def _mejores_peores(cats: dict[str, int], medias: dict[str, float]) -> tuple[list, list]:
    """Top-3 / bottom-3 de categorías por delta contra la media del pool."""
    deltas = [(c, cats.get(c, 0) - medias[c]) for c in medias]
    deltas.sort(key=lambda x: -x[1])
    top = [c for c, d in deltas[:3] if d > 0]
    bottom = [c for c, d in reversed(deltas[-3:]) if d < 0]
    return top, bottom


def build_pool(pool_dir: Path, vps_dir: Path) -> dict:
    admin = pool_dir / "ADMIN.xlsx"
    state = json.load(open(pool_dir / "state.json", encoding="utf-8"))
    players = json.load(open(pool_dir / "players.json", encoding="utf-8"))
    slot_by_name = {p["name"]: p["slot"] for p in players}

    # jid rescatado del VPS (puede no existir el fichero o el jugador)
    jid_by_name: dict[str, str] = {}
    vps_players = vps_dir / "players.json"
    if vps_players.exists():
        for p in json.load(open(vps_players, encoding="utf-8")):
            if p.get("whatsapp_jid"):
                jid_by_name[p["name"]] = p["whatsapp_jid"]

    ranking = read_ranking(admin)
    breakdown = read_clas_breakdown(admin)
    honor = _read_honor_block(admin, slot_by_name)
    ex_cnt, sg_cnt = _exactos_signos(state.get("all_predictions") or [])
    ko = state.get("knockout") or {}
    ko_por_jugador = ko.get("por_jugador") or {}
    ko_picks = ko.get("picks") or {}

    # asserts de consistencia (fallar ruidosamente antes que publicar mentiras)
    lb = {e["name"]: (e["position"], e["points"]) for e in state["leaderboard"]}
    for e in ranking:
        b = breakdown[e["name"]]
        assert sum(b["cats"].values()) == b["total"] == e["points"], \
            f"{e['name']}: desglose {sum(b['cats'].values())} != total {b['total']} != ranking {e['points']}"
        assert lb[e["name"]] == (e["position"], e["points"]), \
            f"{e['name']}: ranking Excel {(e['position'], e['points'])} != state.leaderboard {lb[e['name']]}"

    # medias del pool por categoría (solo categorías con algún punto en el pool)
    activas = [c for c in CLAS_CATEGORIES
               if any(breakdown[n]["cats"].get(c, 0) for n in breakdown)]
    medias = {c: sum(breakdown[n]["cats"].get(c, 0) for n in breakdown) / len(breakdown)
              for c in activas}

    jugadores = {}
    for e in ranking:
        name = e["name"]
        cats = breakdown[name]["cats"]
        top, bottom = _mejores_peores(cats, medias)
        pj = ko_por_jugador.get(name) or {}
        highlights = []
        for cat_label in ("Equipos Final", "Equipos 1/2"):
            aciertos = (pj.get(cat_label) or {}).get("aciertos") or []
            if aciertos:
                highlights.append(f"{cat_label}: acertó {', '.join(aciertos)}")
        perfil = next(p for p in players if p["name"] == name)
        jugadores[name] = {
            "pos": e["position"], "puntos": e["points"],
            "desglose": cats,
            "exactos": ex_cnt.get(name, 0), "signos": sg_cnt.get(name, 0),
            "mejores_categorias": top, "peores_categorias": bottom,
            "ko": {"por_categoria": pj, "picks": ko_picks.get(name) or {},
                   "highlights": highlights},
            "honor": honor[name],
            "perfil": {"club": perfil.get("club"), "national": perfil.get("national")},
            "whatsapp_jid": jid_by_name.get(name),
        }

    return {
        "final_leaderboard": [{"pos": e["position"], "name": e["name"],
                               "points": e["points"]} for e in ranking],
        "players": jugadores,
    }


def render_md(palmares: dict) -> str:
    lines = [f"# Palmarés — {palmares['edition']}", ""]
    r = palmares["real"]
    lines += [f"**Final:** {r['final']} · **3er puesto:** {r['tercer_puesto']}",
              f"**Balón de Oro:** {r['balon_oro']} · **Bota de Oro:** {r['bota_oro']}", ""]
    for nota in palmares["notas"]:
        lines.append(f"- {nota}")
    for pool, data in palmares["pools"].items():
        lines += ["", f"## {pool}", ""]
        for e in data["final_leaderboard"]:
            j = data["players"][e["name"]]
            extras = [f"{j['exactos']} exactos", f"{j['signos']} signos"]
            if j["mejores_categorias"]:
                extras.append("fuerte en " + ", ".join(j["mejores_categorias"]))
            if j["peores_categorias"]:
                extras.append("flojo en " + ", ".join(j["peores_categorias"]))
            honor_hits = [k for k, v in j["honor"].items() if v["puntos"] > 0]
            if honor_hits:
                extras.append("honor: " + ", ".join(honor_hits))
            for h in j["ko"]["highlights"]:
                extras.append(h)
            lines.append(f"{e['pos']}. **{e['name']}** — {e['points']} pts "
                         f"({'; '.join(extras)})")
    return "\n".join(lines) + "\n"


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--pools-dir", default="editions/mundial-2026/pools")
    ap.add_argument("--vps-dir", default="editions/mundial-2026/vps")
    ap.add_argument("--out-dir", default="editions/mundial-2026")
    ap.add_argument("--edition", default="mundial-2026")
    ap.add_argument("--meta", default="editions/mundial-2026/palmares_meta.json",
                    help='JSON {"real": {...}, "notas": [...]} de la edición')
    args = ap.parse_args()

    pools_dir = PROJECT_ROOT / args.pools_dir
    vps_dir = PROJECT_ROOT / args.vps_dir
    out_dir = PROJECT_ROOT / args.out_dir
    meta = json.load(open(PROJECT_ROOT / args.meta, encoding="utf-8"))

    palmares = {"edition": args.edition, "real": meta["real"],
                "notas": meta["notas"], "pools": {}}
    for pool_dir in sorted(p for p in pools_dir.iterdir() if p.is_dir()):
        pool = pool_dir.name
        palmares["pools"][pool] = build_pool(pool_dir, vps_dir / pool)
        n = len(palmares["pools"][pool]["players"])
        print(f"[{pool}] {n} jugadores destilados")

    out_json = out_dir / "palmares.json"
    out_json.write_text(json.dumps(palmares, ensure_ascii=False, indent=2),
                        encoding="utf-8")
    out_md = out_dir / "palmares.md"
    out_md.write_text(render_md(palmares), encoding="utf-8")
    print(f"OK -> {out_json}\nOK -> {out_md}")


if __name__ == "__main__":
    main()
