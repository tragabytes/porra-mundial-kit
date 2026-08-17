"""Tests del re-cálculo independiente de clasificados (módulo 10 / lib_scoring).

No tocan Excels reales: usan workbooks sintéticos en tmp y datos en memoria.
"""
from openpyxl import Workbook

import lib_scoring as sc
from set_scoring import SCORING


# --- Baremo ----------------------------------------------------------------

def test_baremo_desde_set_scoring():
    b = sc.knockout_equipos_baremo()
    assert b["Equipos 1/16"] == SCORING[15] == 1
    assert b["Equipos 1/8"] == SCORING[19] == 2
    assert b["Equipos 1/4"] == SCORING[23] == 3
    assert b["Equipos 1/2"] == SCORING[27] == 4
    assert b["Equipos 3-4"] == SCORING[31] == 4
    assert b["Equipos Final"] == SCORING[32] == 6


# --- Clasificados reales desde la API --------------------------------------

def test_actuals_from_api():
    api = [
        {"stage": "LAST_32", "home": "Spain", "away": "Egypt"},
        {"stage": "LAST_32", "home": "Brazil", "away": "Japan"},
        {"stage": "LAST_16", "home": None, "away": None},   # cruce sin resolver
        {"stage": "GROUP_STAGE", "home": "Spain", "away": "X"},  # se ignora
    ]
    es = {"Spain": "España", "Egypt": "Egipto", "Brazil": "Brasil", "Japan": "Japón"}
    out = sc.knockout_actuals_from_api(api, es)
    assert out["Equipos 1/16"] == {"España", "Egipto", "Brasil", "Japón"}
    assert out["Equipos 1/8"] == set()     # ronda sin resolver
    assert out["Equipos 1/4"] == set()


# --- Re-cálculo: aciertos / fallos / puntos / resolución parcial ------------

def test_recompute_ronda_resuelta():
    picks = {"Ana": {"Equipos 1/16": ["Brasil", "Francia", "Alemania"]}}
    actual = {"Equipos 1/16": {"Brasil", "Francia", "Egipto"}}
    baremo = {"Equipos 1/16": 1}
    r = sc.knockout_qualifier_recompute(picks, actual, baremo)["Ana"]["Equipos 1/16"]
    assert r["n"] == 2
    assert r["puntos"] == 2
    assert r["aciertos"] == ["Brasil", "Francia"]
    assert r["fallos"] == ["Alemania"]      # Egipto no estaba en sus picks
    assert r["resuelta"] is True


def test_recompute_baremo_mayor():
    picks = {"Ana": {"Equipos 1/4": ["Brasil", "Francia"]}}
    actual = {"Equipos 1/4": {"Brasil", "Francia"}}
    r = sc.knockout_qualifier_recompute(picks, actual, {"Equipos 1/4": 3})["Ana"]["Equipos 1/4"]
    assert r["n"] == 2 and r["puntos"] == 6  # 2 aciertos * 3


def test_recompute_ronda_no_resuelta_no_marca_fallos():
    picks = {"Ana": {"Equipos 1/8": ["Brasil", "Francia"]}}
    actual = {"Equipos 1/8": set()}          # aún no resuelta
    r = sc.knockout_qualifier_recompute(picks, actual, {"Equipos 1/8": 2})["Ana"]["Equipos 1/8"]
    assert r["n"] == 0 and r["puntos"] == 0
    assert r["fallos"] == []                  # sin resolver: un equipo no es fallo todavía
    assert r["resuelta"] is False


# --- Lector de picks desde el ADMIN ----------------------------------------

def _make_admin(tmp_path, players):
    """players: lista de (slot, nombre, {categoria: [equipos]})."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ADMIN"
    ws["D5"] = max(s for s, *_ in players)
    for slot, name, cats in players:
        col = 19 + (slot - 1) * 3            # S, V, Y…
        ws.cell(row=5, column=col).value = name
        for cat, r0, r1, _ in sc.KO_EQUIPOS_BLOCKS:
            for i, team in enumerate(cats.get(cat, [])):
                ws.cell(row=r0 + i, column=col).value = team
    p = tmp_path / "ADMIN.xlsx"
    wb.save(str(p))
    return p


def test_read_qualifier_picks(tmp_path):
    admin = _make_admin(tmp_path, [
        (1, "ROBER", {"Equipos 1/16": ["Brasil", "Francia"], "Equipos 1/8": ["Brasil"]}),
        (2, "Nico", {"Equipos 1/16": ["España"]}),
    ])
    picks = sc.read_knockout_qualifier_picks(admin)
    assert picks["ROBER"]["Equipos 1/16"] == ["Brasil", "Francia"]
    assert picks["ROBER"]["Equipos 1/8"] == ["Brasil"]
    assert picks["Nico"]["Equipos 1/16"] == ["España"]
    assert picks["Nico"]["Equipos 1/8"] == []   # vacío, no presente


def test_read_qualifier_picks_ignora_slots_sin_cargar(tmp_path):
    admin = _make_admin(tmp_path, [(1, "Pegar Valores Nombre J1", {})])
    assert sc.read_knockout_qualifier_picks(admin) == {}


def test_end_to_end_recompute_cuadra(tmp_path):
    admin = _make_admin(tmp_path, [
        (1, "ROBER", {"Equipos 1/16": ["Brasil", "Francia", "Alemania"]}),
    ])
    picks = sc.read_knockout_qualifier_picks(admin)
    api = [{"stage": "LAST_32", "home": "Brazil", "away": "France"}]
    es = {"Brazil": "Brasil", "France": "Francia"}
    actual = sc.knockout_actuals_from_api(api, es)
    r = sc.knockout_qualifier_recompute(picks, actual, sc.knockout_equipos_baremo())
    assert r["ROBER"]["Equipos 1/16"]["n"] == 2   # Brasil + Francia (Alemania no está)
    assert r["ROBER"]["Equipos 1/16"]["puntos"] == 2


# --- Marcador KO fiel al cruce (predicción por partido en eliminatorias) -----

def test_stage_y_baremo_exacto():
    assert sc.KO_STAGE_TO_RONDA["LAST_32"] == "Dieciseisavos"
    assert sc.KO_STAGE_TO_RONDA["FINAL"] == "Final"
    # signo + exacto por ronda
    assert sc.KO_RONDA_EXACTO_PTS["Dieciseisavos"] == 6
    assert sc.KO_RONDA_EXACTO_PTS["Cuartos"] == 8
    assert sc.KO_RONDA_EXACTO_PTS["Final"] == 13


def test_ronda_for_admin_row():
    assert sc.ronda_for_admin_row(164) == "Dieciseisavos"
    assert sc.ronda_for_admin_row(179) == "Dieciseisavos"
    assert sc.ronda_for_admin_row(200) == "Octavos"
    assert sc.ronda_for_admin_row(247) == "Final"
    assert sc.ronda_for_admin_row(180) is None   # hueco entre bloques
    assert sc.ronda_for_admin_row(38) is None     # fila de fase de grupos


_MATCHUPS = {
    # orientación normal
    "ROBER": {"Dieciseisavos": [{"cruce": "España-Italia", "marcador": "2-1"}]},
    # orientación invertida (escribió Italia-España, marcador en su orden)
    "Nico": {"Dieciseisavos": [{"cruce": "Italia-España", "marcador": "1-2"}]},
    # otro cruce distinto: no debe aparecer
    "Ana": {"Dieciseisavos": [{"cruce": "Brasil-Japón", "marcador": "3-0"}]},
}


def test_marcadores_orientacion_normal_e_invertida_exacto():
    out = sc.knockout_marcadores(_MATCHUPS, "Dieciseisavos", "España", "Italia", 2, 1)
    assert set(out) == {"ROBER", "Nico"}          # Ana (otro cruce) fuera
    assert out["ROBER"] == {"marcador": "2-1", "exacto": True}
    # Nico escribió Italia-España 1-2 = España 2 - Italia 1 → se normaliza a 2-1
    assert out["Nico"] == {"marcador": "2-1", "exacto": True}


def test_marcadores_no_exacto():
    out = sc.knockout_marcadores(_MATCHUPS, "Dieciseisavos", "España", "Italia", 3, 1)
    assert out["ROBER"]["exacto"] is False
    assert out["ROBER"]["marcador"] == "2-1"        # su predicción, intacta


def test_marcadores_sin_resultado_es_None():
    out = sc.knockout_marcadores(_MATCHUPS, "Dieciseisavos", "España", "Italia")
    assert out["ROBER"]["exacto"] is None
    assert out["Nico"]["marcador"] == "2-1"


def test_marcadores_cruce_no_predicho_vacio():
    out = sc.knockout_marcadores(_MATCHUPS, "Dieciseisavos", "Portugal", "Uruguay", 0, 0)
    assert out == {}


def test_marcadores_equipo_con_guion_no_se_trocea():
    # El cruce se compara como string completo (hay países con guion).
    mu = {"Ann": {"Octavos": [{"cruce": "Bosnia-Herzegovina-Croacia", "marcador": "1-0"}]}}
    out = sc.knockout_marcadores(mu, "Octavos", "Bosnia-Herzegovina", "Croacia", 1, 0)
    assert out["Ann"] == {"marcador": "1-0", "exacto": True}


def test_predicciones_formato_tupla_con_signo():
    out = sc.knockout_predicciones(_MATCHUPS, "Dieciseisavos", "España", "Italia")
    assert out["ROBER"] == ("1", 2, 1)              # 2-1 → gana local
    assert out["Nico"] == ("1", 2, 1)           # normalizado a orientación real
    assert "Ana" not in out


# --- Clasificación por cruce (cruce / ambos / un equipo / nada) --------------

_CLASIF = {
    # acertó el cruce exacto (orientación normal)
    "ROBER": {"Dieciseisavos": [{"cruce": "España-Italia", "marcador": "2-1"}]},
    # acertó el cruce (orientación invertida); marcador 0-0 → no exacto vs 2-1
    "Nico": {"Dieciseisavos": [{"cruce": "Italia-España", "marcador": "0-0"}]},
    # solo el local (España con otro rival)
    "Ana": {"Dieciseisavos": [{"cruce": "España-Portugal", "marcador": "1-0"}]},
    # solo el visitante (Italia con otro rival)
    "Leo": {"Dieciseisavos": [{"cruce": "Brasil-Italia", "marcador": "2-2"}]},
    # los dos equipos pero en cruces distintos (sin emparejar entre sí)
    "Bea": {"Dieciseisavos": [{"cruce": "España-Francia", "marcador": "1-0"},
                              {"cruce": "Italia-Japón", "marcador": "3-1"}]},
    # ninguno de los dos
    "Tom": {"Dieciseisavos": [{"cruce": "Brasil-Japón", "marcador": "3-0"}]},
    # tiene picks en otra ronda, pero Dieciseisavos vacío → nada
    "Eva": {"Octavos": [{"cruce": "España-Italia", "marcador": "2-1"}]},
}


def test_clasificacion_cuatro_grupos():
    c = sc.knockout_clasificacion(_CLASIF, "Dieciseisavos", "España", "Italia", 2, 1)
    assert set(c["cruce"]) == {"ROBER", "Nico"}
    assert c["cruce"]["ROBER"] == {"marcador": "2-1", "signo": "1", "exacto": True}
    assert c["cruce"]["Nico"]["exacto"] is False     # acertó cruce, no marcador
    assert c["ambos"] == ["Bea"]
    assert c["un_equipo"]["Ana"] == "España"
    assert c["un_equipo"]["Leo"] == "Italia"
    assert set(c["nada"]) == {"Tom", "Eva"}            # otro cruce + ronda vacía


def test_clasificacion_sin_resultado_exacto_none():
    c = sc.knockout_clasificacion(_CLASIF, "Dieciseisavos", "España", "Italia")
    assert c["cruce"]["ROBER"]["exacto"] is None
    assert c["cruce"]["ROBER"]["signo"] == "1"


def test_clasificacion_equipo_con_guion():
    mu = {
        "Ann": {"Octavos": [{"cruce": "Bosnia-Herzegovina-Croacia", "marcador": "1-0"}]},
        "Bob": {"Octavos": [{"cruce": "Croacia-Italia", "marcador": "2-1"}]},  # solo Croacia
    }
    c = sc.knockout_clasificacion(mu, "Octavos", "Bosnia-Herzegovina", "Croacia", 1, 0)
    assert "Ann" in c["cruce"]
    assert c["un_equipo"]["Bob"] == "Croacia"


def test_clasificacion_regresion_hugo():
    # Hugo predijo el cruce "Corea del Sur-Canadá" 0-1; el partido real fue
    # Sudáfrica 0-1 Canadá: acertó solo a Canadá, NO el cruce → no clava.
    mu = {"Hugo": {"Dieciseisavos": [{"cruce": "Corea del Sur-Canadá", "marcador": "0-1"}]}}
    c = sc.knockout_clasificacion(mu, "Dieciseisavos", "Sudáfrica", "Canadá", 0, 1)
    assert c["cruce"] == {}                             # nadie acertó el cruce
    assert c["un_equipo"]["Hugo"] == "Canadá"
    assert [n for n, v in c["cruce"].items() if v["exacto"]] == []


# --- Puntos por partido KO (itemización de la web, no "+0") -----------------

def test_ko_signo_pts_desde_set_scoring():
    assert sc.KO_RONDA_SIGNO_PTS["Dieciseisavos"] == SCORING[16] == 2
    assert sc.KO_RONDA_SIGNO_PTS["Octavos"] == SCORING[20] == 2
    assert sc.KO_RONDA_SIGNO_PTS["Cuartos"] == SCORING[24] == 3
    assert sc.KO_RONDA_SIGNO_PTS["Semifinales"] == SCORING[28] == 4
    assert sc.KO_RONDA_SIGNO_PTS["Final"] == SCORING[36] == 5


def test_ko_match_points_exacto():
    # Predijo el cruce y clavó 2-1 en octavos -> signo + exacto.
    assert sc.knockout_match_points(("1", 2, 1), 2, 1, "Octavos") == sc.KO_RONDA_EXACTO_PTS["Octavos"]


def test_ko_match_points_solo_signo():
    # Predijo 3-0 (local); salió 2-1 (local) -> solo signo, no exacto.
    assert sc.knockout_match_points(("1", 3, 0), 2, 1, "Octavos") == sc.KO_RONDA_SIGNO_PTS["Octavos"]


def test_ko_match_points_fallo():
    # Predijo victoria local; salió visitante -> 0.
    assert sc.knockout_match_points(("1", 2, 1), 0, 1, "Octavos") == 0


def test_ko_itemizacion_como_build_web_data():
    # Reproduce el caso Marta: preds solo trae a quien predijo ESTE cruce, y la
    # web itemiza los puntos con knockout_match_points (antes se dejaban en +0).
    # Marta predijo Portugal-España 2-3 (signo 2); el real fue Portugal 0-1 España
    # (signo 2) → acertó el signo: +2 en octavos, NO +0. Curro predijo otro cruce.
    mu = {
        "Marta": {"Octavos": [{"cruce": "Portugal-España", "marcador": "2-3"}]},
        "Curro": {"Octavos": [{"cruce": "Francia-Brasil", "marcador": "1-0"}]},
    }
    preds = sc.knockout_predicciones(mu, "Octavos", "Portugal", "España")
    assert set(preds) == {"Marta"}                     # Curro no predijo este cruce → fuera
    pts = {n: sc.knockout_match_points(p, 0, 1, "Octavos") for n, p in preds.items()}
    assert pts["Marta"] == sc.KO_RONDA_SIGNO_PTS["Octavos"] == 2
