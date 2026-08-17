"""Tests de la auditoría: scoring independiente, escáner de errores y archivado.

No tocan Excels reales: usan workbooks sintéticos en tmp y monkeypatch.
"""
import json

from openpyxl import Workbook

import lib_scoring as sc
import lib_whatsapp_client as wac
import audit_pool


# --- Informe JSON de auditoría (Section.to_json) ---------------------------

def test_section_to_json():
    s = audit_pool.Section("1. Prueba", id="prueba", que="qué se mira", como="cómo se mira")
    s.ok("todo correcto")
    s.warn("una pega")
    s.evid({"dato": 1})
    j = s.to_json()
    assert j["id"] == "prueba"
    assert j["que"] == "qué se mira"
    assert j["como"] == "cómo se mira"
    assert j["status"] == audit_pool.WARN          # un warn sube el estado
    assert j["resultado"] == "todo correcto"        # las líneas OK
    assert j["hallazgos"] == [{"status": audit_pool.WARN, "msg": "una pega"}]
    assert j["evidencia"] == [{"dato": 1}]


# --- Re-cálculo de puntos de grupos (debe coincidir con el baremo del Excel) ---

def gp(sign, ph, pa, rh, ra):
    return sc.grupos_match_points((sign, ph, pa), rh, ra)


def test_grupos_points_exacto_signo_fallo():
    assert gp("1", 2, 1, 2, 1) == 4   # exacto
    assert gp("1", 3, 0, 2, 1) == 1   # solo signo (local gana)
    assert gp("1", 2, 1, 0, 2) == 0   # falla el signo
    assert gp("X", 1, 1, 1, 1) == 4   # empate exacto
    assert gp("X", 0, 0, 1, 1) == 1   # empate, signo
    assert gp("2", 0, 3, 1, 2) == 1   # visitante, signo


def test_grupos_points_coincide_con_ingest():
    # La copia ligera de lib_scoring debe dar lo mismo que la del cron.
    import ingest_match_results as ing
    for sign, ph, pa, rh, ra in [
        ("1", 2, 1, 2, 1), ("1", 3, 0, 2, 1), ("X", 0, 0, 1, 1),
        ("2", 0, 3, 1, 2), ("1", 2, 1, 0, 2), ("X", 1, 1, 1, 1),
    ]:
        assert (sc.grupos_match_points((sign, ph, pa), rh, ra)
                == ing.grupos_match_points((sign, ph, pa), rh, ra))


# --- Escáner de celdas de error -------------------------------------------

def test_find_error_cells_detecta_y_excluye(tmp_path):
    wb = Workbook()
    clas = wb.active
    clas.title = "CLAS"
    clas["A1"] = "ok"
    clas["B2"] = "#REF!"          # debe detectarse
    stats = wb.create_sheet("Stats")
    stats["N5"] = "#NAME?"        # columna N excluida por KNOWN_BROKEN
    stats["A1"] = "#VALUE!"       # NO excluida -> debe detectarse
    path = tmp_path / "wb.xlsx"
    wb.save(path)

    exclude = {"Stats": {"N"}}
    bad = sc.find_error_cells(path, ["CLAS", "Stats"], exclude)
    joined = " ".join(bad)
    assert "CLAS!B2" in joined
    assert "Stats!A1" in joined
    assert "Stats!N5" not in joined   # excluida por columna


# --- Desglose por categoría de CLAS (consistencia total == suma) -----------

def test_read_clas_breakdown(tmp_path):
    wb = Workbook()
    adm = wb.active
    adm.title = "ADMIN"
    adm["D5"] = 2  # 2 jugadores
    clas = wb.create_sheet("CLAS")
    # Ana: total 10 = F.Grupos 7 (E) + Pos.Grupos 3 (F)
    clas.cell(row=5, column=3).value = "Ana"
    clas.cell(row=5, column=4).value = 10
    clas.cell(row=5, column=5).value = 7
    clas.cell(row=5, column=6).value = 3
    # Beto: total 5 = F.Grupos 5
    clas.cell(row=6, column=3).value = "Beto"
    clas.cell(row=6, column=4).value = 5
    clas.cell(row=6, column=5).value = 5
    path = tmp_path / "clas.xlsx"
    wb.save(path)

    bd = sc.read_clas_breakdown(path)
    assert bd["Ana"]["total"] == 10
    assert bd["Ana"]["cats"]["F. Grupos"] == 7
    assert bd["Ana"]["cats"]["Pos. Grupos"] == 3
    assert sum(bd["Ana"]["cats"].values()) == bd["Ana"]["total"]   # consistencia
    assert bd["Beto"]["total"] == 5
    assert sum(bd["Beto"]["cats"].values()) == 5


# --- Archivado de mensajes enviados ---------------------------------------

def test_archive_sent_message(tmp_path, monkeypatch):
    monkeypatch.setattr(wac, "PROJECT_ROOT", tmp_path)
    monkeypatch.setenv("POOL_ID", "testpool")
    wac._archive_sent_message("⚽💥 hola", "34600@g.us", has_image=True)
    log = tmp_path / "pools" / "testpool" / "sent_messages.jsonl"
    assert log.exists()
    rec = json.loads(log.read_text(encoding="utf-8").strip())
    assert rec["pool"] == "testpool"
    assert rec["text"] == "⚽💥 hola"
    assert rec["group_id"] == "34600@g.us"
    assert rec["image"] is True
    assert rec["chars"] == len("⚽💥 hola")


def test_archive_sin_pool_no_escribe(tmp_path, monkeypatch):
    monkeypatch.setattr(wac, "PROJECT_ROOT", tmp_path)
    monkeypatch.delenv("POOL_ID", raising=False)
    wac._archive_sent_message("hola", "g@g.us", has_image=False)
    assert not (tmp_path / "pools").exists()
